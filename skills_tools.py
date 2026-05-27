from __future__ import annotations

import io
import re
from datetime import datetime

import aiohttp
import edge_tts
from aiogram import Bot, Router
from aiogram.filters import Command
from aiogram.types import FSInputFile, Message
from bs4 import BeautifulSoup
from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

from config import settings
from telegram_format import reply_formatted

router = Router(name="skills_tools")

_URL_RE = re.compile(r"(https?://[^\s]+)", re.IGNORECASE)


def _safe_filename(prefix: str, ext: str) -> str:
    ts = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    return f"{prefix}-{ts}.{ext.lstrip('.')}"


@router.message(Command("tts"))
async def cmd_tts(message: Message, bot: Bot) -> None:
    text = (message.text or "").strip()
    text = text.removeprefix("/tts").strip()
    if not text:
        await reply_formatted(
            message,
            "🎙️ **/tts** текст\n\n"
            "🔹 Пример: `/tts Добрый вечер, чат`",
        )
        return

    voice = "ru-RU-DmitryNeural"
    out_path = settings.downloads_dir / _safe_filename("tts", "mp3")
    out_path.parent.mkdir(parents=True, exist_ok=True)

    communicate = edge_tts.Communicate(text, voice=voice)
    await communicate.save(str(out_path))

    await message.reply_voice(FSInputFile(out_path))


@router.message(Command("scrape"))
async def cmd_scrape(message: Message) -> None:
    text = (message.text or "").strip()
    text = text.removeprefix("/scrape").strip()
    m = _URL_RE.search(text)
    if not m:
        await reply_formatted(
            message,
            "🕸️ **/scrape** ссылка\n\n"
            "🔹 Пример: `/scrape https://example.com`",
        )
        return
    url = m.group(1).rstrip(").,]")

    async with aiohttp.ClientSession() as session:
        async with session.get(url, timeout=aiohttp.ClientTimeout(total=15)) as resp:
            html = await resp.text(errors="ignore")

    soup = BeautifulSoup(html, "html.parser")
    title = (soup.title.string or "").strip() if soup.title and soup.title.string else ""
    h1 = soup.find("h1")
    h1_text = h1.get_text(" ", strip=True) if h1 else ""

    out = f"🕸️ **Снял шапку страницы**\n\n🔗 {url}"
    if title:
        out += f"\n\n🏷️ **Title**: {title}"
    if h1_text and h1_text != title:
        out += f"\n\n🧾 **H1**: {h1_text}"

    await reply_formatted(message, out)


@router.message(Command("pdf"))
async def cmd_pdf(message: Message, bot: Bot) -> None:
    if not message.reply_to_message or not message.reply_to_message.document:
        await reply_formatted(
            message,
            "📄 **/pdf** — ответь этой командой на PDF-файл.\n\n"
            "🔹 Я вытащу текст и дам короткую выжимку.",
        )
        return

    doc = message.reply_to_message.document
    if not (doc.mime_type or "").lower().endswith("pdf") and not doc.file_name.lower().endswith(".pdf"):
        await reply_formatted(message, "📄 Это не похоже на **PDF**.")
        return

    buf = io.BytesIO()
    await bot.download(doc.file_id, destination=buf)
    buf.seek(0)

    reader = PdfReader(buf)
    chunks: list[str] = []
    for page in reader.pages[:10]:
        t = (page.extract_text() or "").strip()
        if t:
            chunks.append(t)
    text = "\n".join(chunks).strip()

    if not text:
        await reply_formatted(message, "📄 В PDF не нашёл читаемого текста (возможно, это сканы).")
        return

    snippet = text.replace("\n", " ")
    if len(snippet) > 1200:
        snippet = snippet[:1200] + "…"
    await reply_formatted(
        message,
        "📄 **PDF → текст**\n\n"
        f"🧾 {snippet}",
    )


@router.message(Command("docx"))
async def cmd_docx(message: Message, bot: Bot) -> None:
    if not message.reply_to_message or not message.reply_to_message.document:
        await reply_formatted(
            message,
            "📝 **/docx** — ответь этой командой на DOCX-файл.\n\n"
            "🔹 Я вытащу текст и покажу кусок.",
        )
        return

    doc = message.reply_to_message.document
    if not doc.file_name.lower().endswith(".docx"):
        await reply_formatted(message, "📝 Это не похоже на **DOCX**.")
        return

    buf = io.BytesIO()
    await bot.download(doc.file_id, destination=buf)
    buf.seek(0)

    d = Document(buf)
    text = "\n".join(p.text for p in d.paragraphs if p.text).strip()
    if not text:
        await reply_formatted(message, "📝 В DOCX пусто или текст не извлёкся.")
        return

    snippet = text.replace("\n", " ")
    if len(snippet) > 1200:
        snippet = snippet[:1200] + "…"
    await reply_formatted(message, f"📝 **DOCX → текст**\n\n🧾 {snippet}")


@router.message(Command("xlsx"))
async def cmd_xlsx(message: Message, bot: Bot) -> None:
    if not message.reply_to_message or not message.reply_to_message.document:
        await reply_formatted(
            message,
            "📊 **/xlsx** — ответь этой командой на XLSX-файл.\n\n"
            "🔹 Я покажу первые строки первого листа.",
        )
        return

    doc = message.reply_to_message.document
    if not doc.file_name.lower().endswith(".xlsx"):
        await reply_formatted(message, "📊 Это не похоже на **XLSX**.")
        return

    buf = io.BytesIO()
    await bot.download(doc.file_id, destination=buf)
    buf.seek(0)

    wb = load_workbook(buf, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for r_i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_i > 15:
            break
        cells = ["" if v is None else str(v) for v in row[:8]]
        rows.append(" | ".join(cells).strip())

    if not rows:
        await reply_formatted(message, "📊 В XLSX пусто или не смог прочитать.")
        return

    preview = "\n".join(rows)
    await reply_formatted(message, f"📊 **XLSX превью**\n\n🧾 {preview}")

