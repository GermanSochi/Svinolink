from __future__ import annotations

import io
import re

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader


def extract_pdf_text(data: bytes, *, max_pages: int = 10) -> str:
    buf = io.BytesIO(data)
    reader = PdfReader(buf)
    chunks: list[str] = []
    for page in reader.pages[:max_pages]:
        t = (page.extract_text() or "").strip()
        if t:
            chunks.append(t)
    return "\n".join(chunks).strip()


def extract_docx_text(data: bytes) -> str:
    buf = io.BytesIO(data)
    d = Document(buf)
    return "\n".join(p.text for p in d.paragraphs if p.text).strip()


def extract_xlsx_preview(data: bytes, *, max_rows: int = 15, max_cols: int = 8) -> str:
    buf = io.BytesIO(data)
    wb = load_workbook(buf, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[str] = []
    for r_i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_i > max_rows:
            break
        cells = [_clean_cell(v) for v in row[:max_cols]]
        # убираем пустые хвосты
        while cells and not cells[-1]:
            cells.pop()
        if not any(cells):
            continue
        # более читаемый разделитель вместо |
        rows.append(" · ".join(c for c in cells if c))

    if not rows:
        return ""

    # Превью как список строк: каждую строку пометим эмодзи-якорем.
    out_lines = []
    marks = ("📌", "🔹", "🧾", "🧩", "⚙️", "🎯", "🗂️", "🧷", "📝", "📎")
    for i, r in enumerate(rows):
        out_lines.append(f"{marks[i % len(marks)]} {r}")
    return "\n".join(out_lines).strip()


def _clean_cell(v) -> str:
    if v is None:
        return ""
    s = str(v)
    # заменяем "забор" и вертикальные разделители
    s = s.replace("|", " ").replace("│", " ")
    s = re.sub(r"\s+", " ", s).strip()
    # ограничение длины ячейки, чтобы не рвало сообщение
    if len(s) > 60:
        s = s[:60].rstrip() + "…"
    return s


def extract_plain_text(data: bytes) -> str:
    # Пытаемся UTF-8, затем Windows-1251 (часто для русских txt).
    for enc in ("utf-8", "utf-8-sig", "cp1251"):
        try:
            return data.decode(enc).strip()
        except UnicodeDecodeError:
            continue
    # Последний шанс: игнорируем битые символы.
    return data.decode("utf-8", errors="ignore").strip()

